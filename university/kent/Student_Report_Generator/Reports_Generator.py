import argparse
import json
import os
import zipfile

import pandas as pd
from docx import Document
from openpyxl import load_workbook

# For debugging:
verbose = False


def parse_arguments():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(description="Generate student reports from an Excel marking file.")
    parser.add_argument("--config", default="config.json", help="Path to the config file.")
    return parser.parse_args()


def load_config(config_path):
    """Load configuration from JSON file."""
    with open(config_path, "r") as f:
        return json.load(f)


def clean_output_folder(output_folder):
    """Remove all existing .xlsx files from the output folder."""
    if os.path.exists(output_folder):
        for file in os.listdir(output_folder):
            if file.endswith(".xlsx"):
                os.remove(os.path.join(output_folder, file))
        print(f"🧹 Cleaned output folder: {output_folder}")


def validate_row_limits(start_row, end_row, total_rows):
    """Validate and adjust start_row and end_row values."""
    if not isinstance(start_row, int) or start_row < 1:
        print(f"⚠️ Warning: Invalid start_row ({start_row}) in config. Defaulting to 1.")
        start_row = 1
    elif start_row > total_rows:
        print(f"⚠️ Warning: start_row ({start_row}) exceeds total rows ({total_rows}). Defaulting to 1.")
        start_row = 1

    if end_row is not None:
        if not isinstance(end_row, int) or end_row < start_row:
            print(f"⚠️ Warning: Invalid end_row ({end_row}) in config. Defaulting to total rows ({total_rows}).")
            end_row = total_rows
        else:
            end_row = min(end_row, total_rows)

    return start_row - 1, end_row  # Convert to 0-based index for Pandas


def load_dataframe(filename, sheet_name=0):
    """Load marking data from a CSV or XLSX file into a DataFrame."""
    file_extension = os.path.splitext(filename)[-1].lower()
    if file_extension == ".csv":
        df = pd.read_csv(filename)
    elif file_extension in [".xls", ".xlsx"]:
        df = pd.read_excel(filename, sheet_name=sheet_name)
    else:
        raise ValueError(f"Unsupported file format: {file_extension}")

    # Ensure column names are consistent
    df.rename(columns=str.strip, inplace=True)

    return df


def load_marking_sheet(marking_file, sheet_name):
    df_marking = load_dataframe(marking_file, sheet_name)

    # Ensure emails are all lower case:
    df_marking["Login"] = df_marking["Login"].str.lower()

    return df_marking


def load_reference_worksheet(reference_file):
    df_reference = load_dataframe(reference_file)

    """Preprocess data: normalize login values and identify missing students."""
    df_reference["Login"] = df_reference["Email address"].str.replace("@kent.ac.uk", "", regex=False)
    df_reference["Login"] = df_reference["Login"].str.lower()

    return df_reference


def validate_reference_worksheet(df_ref, df_marks, module_name, assignment_name):
    # Identify students in df_marking but not in df_reference:
    missing_students = df_marks[~df_marks["Login"].isin(df_ref["Login"])]

    if not missing_students.empty:
        print("⚠️ The following students are in the marking file but missing from the reference file:")
        print(missing_students["Login"].to_string(index=False))

    # Keep the common rows:
    df_ref = df_ref[df_ref["Login"].isin(df_marks["Login"])].copy()

    # Create "Submission_ID" by extracting the second part of the "Identifier" column
    df_ref["Submission_ID"] = df_ref["Identifier"].apply(
        lambda x: str(x).split(" ")[1] if isinstance(x, str) and " " in x else ""
    )

    # Create "submission_file_name" using name + submission_id + module_name + assignment_name
    # Example: (Emad Mahayni_117863_assignsubmission_file_COMPXXXX_A2_Feedback)
    sub_str = f"_assignsubmission_file_{module_name}_{assignment_name}_Feedback - "
    df_ref["feedback_filename"] = df_ref["Full name"] + "_" + df_ref["Submission_ID"] + sub_str + df_ref["Login"]

    return df_ref


def update_reference_file(df_marking, df_reference, marking_workflow_state, output_file):
    """Update the reference file with grading and marking workflow state."""

    df_reference["Marking workflow state"] = marking_workflow_state
    df_reference["Grade"] = df_reference["Login"].map(df_marking.set_index("Login")["Grade"])

    # Remove unnecessary columns
    df_reference.drop(columns=["Login", "Submission_ID", "feedback_filename"], inplace=True)

    """Save the updated reference file."""
    # df_reference.to_excel(output_file, index=False)
    df_reference.to_csv(output_file, index=False)

    print(f"✅ Updated reference file saved as: {output_file}")


def generate_reports(df_marking, config):
    """Generate student reports based on the template and marking file."""

    # Check the feedback template:
    template_file = config["feedback_template_filename"]
    _, ext = os.path.splitext(template_file)
    print(f"Report Template is a {ext} file.")

    output_folder = config["output_folder"]
    key_column = config["key_column"]
    start_row = config.get("start_row", 1)
    end_row = config.get("end_row")
    mapping = config["mapping"]

    os.makedirs(output_folder, exist_ok=True)

    total_rows = len(df_marking)
    start_idx, end_idx = validate_row_limits(start_row, end_row, total_rows)
    generated_count = 0

    for index, row in df_marking.iloc[start_idx:end_idx].iterrows():
        if verbose: print(f"index: {index}, row{row}")

        student_key = row.get(key_column)
        if pd.isna(student_key) or not student_key:
            print(f"Skipping row {index + 1}: Missing key column ({key_column}).")
            continue

        # Create the
        feedback_filename = row.get("feedback_filename")
        output_filename = os.path.join(output_folder, feedback_filename + ext)

        match ext:
            case ".docx":
                write_to_word_file(mapping, row, template_file, output_filename)

            case ".xlsx":
                write_to_excel_file(mapping, row, template_file, output_filename)

            case _:
                print(f"Unsupported Template file: extension {ext}")

        generated_count += 1
        print(f'✅ Report saved: {output_filename}')

    print(f'✨ Total generated reports: {generated_count}')


def write_to_excel_file(mapping, row, template_file, output_filename):
    # Load the template Excel document:
    wb = load_workbook(template_file)
    # Select the first/active sheet:
    ws = wb.active
    for column_name, cell_reference in mapping.items():
        if verbose: print(f"column_name: {column_name}, cell_reference{cell_reference}")
        value = row.get(column_name, "N/A")
        ws[cell_reference] = value if not pd.isna(value) else ""

    # Save the file:
    wb.save(output_filename)


def write_to_word_file(mapping, row, template_file, output_filename):
    # Load the template Word document:
    doc = Document(template_file)
    # Tables in the Word document:
    table = doc.tables[0]
    for column_name, cell_reference in mapping.items():
        if verbose: print(f"column_name: {column_name}, cell_reference{cell_reference}")
        value = row.get(column_name, "N/A")
        cell_val = value if not pd.isna(value) else ""

        i, j = excel_ref_to_index(cell_reference)
        cell = table.cell(i, j)
        cell.text = str(cell_val)

    # Save the file:
    doc.save(output_filename)


def excel_ref_to_index(ref):
    col = ord(ref[0].upper()) - ord('A')
    row = int(ref[1:]) - 1
    return row, col


def zip_output_files(output_folder, zip_filepath):
    """Zip generated reports into a single archive."""
    with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file in os.listdir(output_folder):
            zipf.write(os.path.join(output_folder, file), file)
    print(f'📦 Reports zipped successfully: {zip_filepath}')


def main():
    """Main function to execute the script."""
    args = parse_arguments()
    config = load_config(args.config)

    marking_worksheet_config = config["marking_worksheet"]
    marking_filename = marking_worksheet_config["marking_filename"]
    sheet_name = marking_worksheet_config["sheet_name"]
    module_name = marking_worksheet_config["module_name"]
    assignment_name = marking_worksheet_config["assigment_name"]

    # Clean the output folder before generating the new output:
    clean_output_folder(marking_worksheet_config["output_folder"])
    df_marking = load_marking_sheet(marking_filename, sheet_name)
    df_marking["feedback_filename"] = module_name + "-" + assignment_name + "-" + df_marking['Login']

    reference_worksheet_config = config["reference_worksheet"]

    if reference_worksheet_config["prepare_moodle_files"] == 0:
        generate_reports(df_marking, marking_worksheet_config)

    else:
        moodle_folder = reference_worksheet_config["output_folder"]
        marking_workflow_state = reference_worksheet_config["marking_workflow_state"]
        reference_file = os.path.join(moodle_folder, reference_worksheet_config["reference_file"])
        updated_reference_file = os.path.join(moodle_folder, reference_worksheet_config["reference_file_updated"])

        df_reference = load_reference_worksheet(reference_file)
        df_reference = validate_reference_worksheet(df_reference, df_marking, module_name, assignment_name)

        # feedback_filename would be the Moodle submission_filename:
        df_marking["feedback_filename"] = df_marking["Login"].map(df_reference.set_index("Login")["feedback_filename"])
        update_reference_file(df_marking, df_reference, marking_workflow_state, updated_reference_file)

        generate_reports(df_marking, marking_worksheet_config)

        zip_filename = f"{module_name}_{assignment_name}_Reports.zip"
        zip_filepath = os.path.join(reference_worksheet_config["output_folder"], zip_filename)
        zip_output_files(marking_worksheet_config["output_folder"], zip_filepath)

        clean_output_folder(marking_worksheet_config["output_folder"])


if __name__ == "__main__":
    main()
