import argparse
import shutil

from docx import Document
from openpyxl import load_workbook

from marking_utils import *

# For debugging:
verbose = False
default_config_file = "config.json"


def parse_arguments(config_file):
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(description="Generate student reports from an Excel marking file.")
    parser.add_argument("--config", default=default_config_file, help="Path to the config file.")
    return parser.parse_args()


def validate_row_limits(start_row, end_row, total_rows):
    """Validate and adjust start_row and end_row values."""
    if not isinstance(start_row, int) or start_row < 1:
        print(f"⚠️ Warning: Invalid start_row ({start_row}) in config. Defaulting to 1.")
        start_row = 1
    elif start_row > total_rows:
        print(f"⚠️ Warning: start_row ({start_row}) exceeds total rows ({total_rows}). Defaulting to 1.")
        start_row = 1

    if end_row is not None:
        if not isinstance(end_row, int) or end_row < start_row:
            print(f"⚠️ Warning: Invalid end_row ({end_row}) in config. Defaulting to total rows ({total_rows}).")
            end_row = total_rows
        else:
            end_row = min(end_row, total_rows)

    return start_row - 1, end_row  # Convert to 0-based index for Pandas


def load_marking_sheet(marks_file, sheet_name):
    df_marks = load_dataframe(marks_file, sheet_name)

    # Ensure emails are all lower case:
    df_marks['Login'] = df_marks['Login'].str.lower()
    return df_marks


def load_moodle_worksheet(moodle_file):
    df_moodle = load_dataframe(moodle_file)

    """Preprocess data: normalize login values and identify missing students."""
    df_moodle["Login"] = df_moodle["Email address"].str.replace("@kent.ac.uk", "", regex=False)
    df_moodle['Login'] = df_moodle['Login'].str.lower()

    return df_moodle


def validate_moodle_worksheet(df_moodle, df_marks, module_name, assignment_name):
    # Identify students in df_marks but not in df_moodle:
    missing_students = df_marks[~df_marks["Login"].isin(df_moodle["Login"])]

    if not missing_students.empty:
        print("⚠️ The following students are in the marking file but missing from the moodle file:")
        print(missing_students["Login"].to_string(index=False))

    # Keep the common rows:
    df_moodle = df_moodle[df_moodle["Login"].isin(df_marks["Login"])].copy()

    # Create "Submission_ID" by extracting the second part of the "Identifier" column
    df_moodle["Submission_ID"] = df_moodle["Identifier"].apply(
        lambda x: str(x).split(" ")[1] if isinstance(x, str) and " " in x else ""
    )

    # Create "submission_file_name" using name + submission_id + module_name + assignment_name
    # Example: (Emad Mahayni_117863_assignsubmission_file_COMPXXXX_A2_Feedback)
    sub_str = f"_assignsubmission_file_{module_name}_{assignment_name}_Feedback - "
    df_moodle["feedback_filename"] = df_moodle["Full name"] + "_" + df_moodle["Submission_ID"] + sub_str + df_moodle[
        "Login"]

    return df_moodle


def update_moodle_file(df_marks, df_moodle, moodle_workflow_state, output_file):
    """Update the moodle file with grading and marking workflow state."""

    df_moodle["Marking workflow state"] = moodle_workflow_state
    df_moodle["Grade"] = df_moodle["Login"].map(df_marks.set_index("Login")["Grade"])

    # Remove unnecessary columns
    df_moodle.drop(columns=["Login", "Submission_ID", "feedback_filename"], inplace=True)

    """Save the updated moodle file."""
    # df_moodle.to_excel(output_file, index=False)
    df_moodle.to_csv(output_file, index=False)

    print(f"✅ Updated moodle file saved as: {output_file}")


def generate_reports(df_marks, config):
    """Generate student reports based on the template and marking file."""

    # Check the feedback template:
    template_file = config["feedback_template_filename"]
    _, ext = os.path.splitext(template_file)
    print(f"Report Template is a {ext} file.")

    # Check rows(start, end):
    start_row = config.get("start_row", 1)
    end_row = config.get("end_row")
    start_idx, end_idx = validate_row_limits(start_row, end_row, len(df_marks))
    df_marks = df_marks.iloc[start_idx:end_idx]

    key_column = config["key_column"]
    mapping = config["mapping"]

    # Create output folder:
    output_folder = config["output_folder"]
    os.makedirs(output_folder, exist_ok=True)

    do_generate_reports(df_marks, ext, key_column, mapping, output_folder, template_file)


def do_generate_reports(df_marks, ext, key_column, mapping, output_folder, template_file):
    generated_count = 0
    for index, row in df_marks.iterrows():
        student_key = row.get(key_column)
        if pd.isna(student_key) or not student_key:
            print(f"Skipping row {index + 1}: Missing key column ({key_column}).")
            continue

        # Create the feedback report:
        feedback_filename = row.get("feedback_filename")
        output_filename = os.path.join(output_folder, feedback_filename + ext)

        match ext:
            case ".docx":
                write_to_word_file(mapping, row, template_file, output_filename)

            case ".xlsx":
                write_to_excel_file(mapping, row, template_file, output_filename)

            case _:
                print(f"Unsupported Template file: extension {ext}")

        generated_count += 1
        print(f'✅ Report saved: {output_filename}')
    print(f'✨ Total generated reports: {generated_count}')


def rename_existing_reports(df, feedback_file_extension, reports_folder):
    """
    Rename reports in the reports folder based on the DataFrame containing login, feedback_filename, and ext columns.
    Each file will be renamed from {login}.{ext} to {feedback_filename}.{ext}.
    """
    # Create output sub-folder:
    renamed_output = reports_folder + '/renamed'
    os.makedirs(renamed_output, exist_ok=True)

    renamed_count = 0
    for _, row in df.iterrows():
        login = row.get("Login")
        new_name = row.get("feedback_filename")

        if not login or not new_name:
            print(f"⚠️ Skipping row with missing values: {row}")
            continue

        old_filename = os.path.join(reports_folder, f"{login}.{feedback_file_extension}")
        new_filename = os.path.join(renamed_output, f"{new_name}.{feedback_file_extension}")

        if os.path.exists(old_filename):
            shutil.copy2(old_filename, new_filename)
            print(f"📁 Copied: {old_filename} → {new_filename}")
            renamed_count += 1
        else:
            print(f"❌ File not found: {old_filename}")

    print(f"✅ Total reports renamed: {renamed_count}")


def write_to_excel_file(mapping, row, template_file, output_filename):
    # Load the template Excel document:
    wb = load_workbook(template_file)
    # Select the first/active sheet:
    ws = wb.active
    for column_name, cell_moodle in mapping.items():
        if verbose: print(f"column_name: {column_name}, cell_moodle{cell_moodle}")
        value = row.get(column_name, "N/A")
        ws[cell_moodle] = value if not pd.isna(value) else ""

    # Save the file:
    wb.save(output_filename)


def write_to_word_file(mapping, row, template_file, output_filename):
    # Load the template Word document:
    doc = Document(template_file)
    # Tables in the Word document:
    table = doc.tables[0]
    for column_name, cell_moodle in mapping.items():
        if verbose: print(f"column_name: {column_name}, cell_moodle{cell_moodle}")
        value = row.get(column_name, "N/A")
        cell_val = value if not pd.isna(value) else ""

        i, j = excel_cell_ref_to_indices(cell_moodle)
        cell = table.cell(i, j)
        cell.text = str(cell_val)

    # Save the file:
    doc.save(output_filename)


def generating_reports(config):
    marks_worksheet_config = config["marks_worksheet"]
    marks_filename = marks_worksheet_config["marks_filename"]
    sheet_name = marks_worksheet_config["sheet_name"]
    module_name = marks_worksheet_config["module_name"]
    assignment_name = marks_worksheet_config["assigment_name"]

    # Clean the output folder before generating the new output:
    clean_output_folder(marks_worksheet_config["output_folder"])
    df_marks = load_marking_sheet(marks_filename, sheet_name)
    df_marks["feedback_filename"] = module_name + "-" + assignment_name + "-" + df_marks['Login']

    moodle_worksheet_config = config["moodle_worksheet"]
    if moodle_worksheet_config["prepare_moodle_files"] == 0:
        generate_reports(df_marks, marks_worksheet_config)

    else:
        moodle_folder = moodle_worksheet_config["output_folder"]
        moodle_workflow_state = moodle_worksheet_config["moodle_workflow_state"]
        moodle_file = os.path.join(moodle_folder, moodle_worksheet_config["moodle_file"])
        updated_moodle_file = os.path.join(moodle_folder, moodle_worksheet_config["moodle_file_updated"])

        df_moodle = load_moodle_worksheet(moodle_file)
        df_moodle = validate_moodle_worksheet(df_moodle, df_marks, module_name, assignment_name)

        # feedback_filename would be the Moodle submission_filename:
        df_marks["feedback_filename"] = df_marks["Login"].map(df_moodle.set_index("Login")["feedback_filename"])
        update_moodle_file(df_marks, df_moodle, moodle_workflow_state, updated_moodle_file)

        # Main action:
        generate_reports(df_marks, marks_worksheet_config)

        zip_filename = f"{module_name}_{assignment_name}_Reports.zip"
        zip_filepath = os.path.join(moodle_worksheet_config["output_folder"], zip_filename)
        zip_output_files(marks_worksheet_config["output_folder"], zip_filepath)

        clean_output_folder(marks_worksheet_config["output_folder"])


def renaming_reports(config):
    moodle_worksheet_config = config["moodle_worksheet"]
    if moodle_worksheet_config["prepare_moodle_files"] == 0:
        return

    # Marks Sheet Config:
    marks_worksheet_config = config["marks_worksheet"]
    marks_filename = marks_worksheet_config["marks_filename"]
    sheet_name = marks_worksheet_config["sheet_name"]
    module_name = marks_worksheet_config["module_name"]
    assignment_name = marks_worksheet_config["assigment_name"]
    feedback_file_extension = marks_worksheet_config["feedback_file_extension"]

    df_marks = load_marking_sheet(marks_filename, sheet_name)
    df_marks["feedback_filename"] = module_name + "-" + assignment_name + "-" + df_marks['Login']

    # Marks Worksheet Config:
    moodle_folder = moodle_worksheet_config["output_folder"]
    moodle_workflow_state = moodle_worksheet_config["moodle_workflow_state"]
    moodle_file = os.path.join(moodle_folder, moodle_worksheet_config["moodle_file"])
    updated_moodle_file = os.path.join(moodle_folder, moodle_worksheet_config["moodle_file_updated"])

    df_moodle = load_moodle_worksheet(moodle_file)
    df_moodle = validate_moodle_worksheet(df_moodle, df_marks, module_name, assignment_name)

    # feedback_filename would be the Moodle submission_filename:
    df_marks["feedback_filename"] = df_marks["Login"].map(df_moodle.set_index("Login")["feedback_filename"])
    update_moodle_file(df_marks, df_moodle, moodle_workflow_state, updated_moodle_file)

    # Main action:
    output_folder = marks_worksheet_config["output_folder"]
    rename_existing_reports(df_marks, feedback_file_extension, output_folder)


def main(action):
    """Main function to execute the script."""
    # args = parse_arguments()
    # config = load_config(args.config)
    match action:
        case 1:
            config = load_config("config_generate_reports.json")
            generating_reports(config)

        case 2:
            config = load_config("config_rename_reports.json")
            renaming_reports(config)

        case _:
            print(f"Unsupported Action: {action}")


if __name__ == "__main__":
    main(2)
