import argparse
import json
import os

import pandas as pd
from openpyxl import load_workbook

# For debugging:
verbose = False


def parse_arguments():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(description="Generate student reports from an Excel master file.")
    parser.add_argument("--config", default="config.json", help="Path to the config file.")
    return parser.parse_args()


def load_config(config_path):
    """Load configuration from JSON file."""
    with open(config_path, "r") as f:
        return json.load(f)


def validate_row_limits(start_row, end_row, total_rows):
    """Validate and adjust start_row and end_row values."""
    if not isinstance(start_row, int) or start_row < 1:
        print(f"⚠️ Warning: Invalid start_row ({start_row}) in config. Defaulting to 1.")
        start_row = 1
    elif start_row > total_rows:
        print(f"⚠️ Warning: start_row ({start_row}) exceeds total rows ({total_rows}). Defaulting to 1.")
        start_row = 1

    if end_row is not None:
        if not isinstance(end_row, int) or end_row < start_row:
            print(f"⚠️ Warning: Invalid end_row ({end_row}) in config. Defaulting to total rows ({total_rows}).")
            end_row = total_rows
        else:
            end_row = min(end_row, total_rows)

    return start_row - 1, end_row  # Convert to 0-based index for Pandas


def generate_reports(df, config):
    """Generate student reports based on the template and master file."""
    master_file = config["master_file"]
    template_file = config["template_file"]
    output_folder = config["output_folder"]
    assignment_name = config["assigment_name"]
    key_column = config["key_column"]
    start_row = config.get("start_row", 1)
    end_row = config.get("end_row")
    mapping = config["mapping"]

    os.makedirs(output_folder, exist_ok=True)

    total_rows = len(df)
    start_idx, end_idx = validate_row_limits(start_row, end_row, total_rows)
    generated_count = 0

    for index, row in df.iloc[start_idx:end_idx].iterrows():
        if verbose: print(f"index: {index}, row{row}")

        student_key = row.get(key_column)
        if pd.isna(student_key) or not student_key:
            print(f"Skipping row {index + 1}: Missing key column ({key_column}).")
            continue

        wb = load_workbook(template_file)
        ws = wb.active

        for column_name, cell_reference in mapping.items():
            if verbose: print(f"column_name: {column_name}, cell_reference{cell_reference}")
            value = row.get(column_name, "N/A")
            ws[cell_reference] = value if not pd.isna(value) else "N/A"

        output_filename = os.path.join(output_folder, f"{assignment_name} - {student_key}.xlsx")
        wb.save(output_filename)

        generated_count += 1
        print(f"✅ Report saved: {output_filename}")

    print(f"✨ Total generated reports: {generated_count}")


def main():
    """Main function to execute the script."""
    args = parse_arguments()
    config = load_config(args.config)

    df = pd.read_excel(config["master_file"], sheet_name=config["sheet_name"])
    generate_reports(df, config)


if __name__ == "__main__":
    main()
