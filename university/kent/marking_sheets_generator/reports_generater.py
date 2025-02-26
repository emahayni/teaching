import json
import os
import pandas as pd
from openpyxl import load_workbook

# For debugging:
verbose = False

# Load the config file
config_file = "config.json"
with open(config_file, "r") as f:
    config = json.load(f)

# Extract configuration values
master_file = config["master_file"]
template_file = config["template_file"]
output_folder = config["output_folder"]
assignment_name = config["assigment_name"]
sheet_name = config["sheet_name"]
key_column = config["key_column"]
start_row = config.get("start_row", 1)  # Default to 1 if not provided
end_row = config.get("end_row")  # Process until the last row if not provided
mapping = config["mapping"]

# Ensure the output folder exists
os.makedirs(output_folder, exist_ok=True)

# Load the master sheet
df = pd.read_excel(master_file, sheet_name=sheet_name)
total_rows = len(df)

# Safeguard start_row
if not isinstance(start_row, int) or start_row < 1:
    print(f"⚠️ Warning: Invalid start_row ({start_row}) in config. Defaulting to 1.")
    start_row = 1
elif start_row > total_rows:
    print(f"⚠️ Warning: start_row ({start_row}) exceeds total rows ({total_rows}). Defaulting to 1.")
    start_row = 1

# Safeguard end_row
if end_row is not None:
    if not isinstance(end_row, int) or end_row < start_row:
        print(f"⚠️ Warning: Invalid end_row ({end_row}) in config. Defaulting to total rows ({total_rows}).")
        end_row = total_rows
    else:
        end_row = min(end_row, total_rows)

# Adjust start_row for Pandas indexing (convert from 1-based to 0-based index)
start_idx = start_row - 1
end_idx = end_row  # Pandas iloc uses exclusive end indexing, so this is fine

# Counter for generated reports
generated_count = 0

# Iterate through the specified range of students
for index, row in df.iloc[start_idx:end_idx].iterrows():
    if verbose: print(f"index: {index}, row{row}")

    # Check if key_column exists and is not empty
    student_key = row.get(key_column)
    if pd.isna(student_key) or not student_key:
        print(f"Skipping row {index + 1}: Missing key column ({key_column}).")
        continue

    # Load a fresh copy of the template for each student
    wb = load_workbook(template_file)
    ws = wb.active  # Assuming the data is on the first sheet

    # The data from the master file will be copied to the template based on cell_references in config mapping
    for column_name, cell_reference in mapping.items():
        if verbose: print(f"column_name: {column_name}, cell_reference{cell_reference}")
        value = row.get(column_name, "N/A")  # Use "N/A" if missing
        ws[cell_reference] = value if not pd.isna(value) else "N/A"

    # Save the new report
    output_filename = os.path.join(output_folder, f"{assignment_name} - {student_key}.xlsx")
    wb.save(output_filename)

    generated_count += 1
    print(f"✅ Report saved: {output_filename}")

print(f"✨ Total generated reports: {generated_count}")